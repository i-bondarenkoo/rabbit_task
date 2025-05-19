from abc import ABC, abstractmethod
from openpyxl import Workbook
import csv
import json
import xml.etree.ElementTree as ET


class Report(ABC):
    @abstractmethod
    def generate(self, headers, data, filename):
        pass


class ExcelReport(Report):
    def generate(self, headers: list, data: list, filename: str):
        # создадим объект, новый лист в excel
        wb = Workbook()
        # создадим лист(страницу)
        new_list = wb.active
        new_list.title = "Отчет"
        # запишем имена заголовков в первую строку
        for column_number, header in enumerate(headers, start=1):
            new_list.cell(row=1, column=column_number, value=header)
        # перебираем строку данных
        for row_number, row_data in enumerate(data, start=2):
            for column_number, value in enumerate(row_data, start=1):
                new_list.cell(row=row_number, column=column_number, value=value)
        wb.save(filename)


class CSVReport(Report):
    def generate(self, headers: list, data: list, filename: str):
        # имя файла, режим открытия, кодировка
        with open(
            file="report.csv", mode="w", encoding="utf-8-sig", newline=""
        ) as file:
            writer = csv.writer(file, delimiter=",")
            writer.writerow(headers)
            for row in data:
                writer.writerow(row)


class JSONReport(Report):
    def generate(self, headers: list, data: list, filename: str):
        report_data = []
        with open(file="report.json", mode="w", encoding="utf-8") as file:
            for row in data:
                data_dict = {}
                data_dict = dict(zip(headers, row))
                report_data.append(data_dict)
            report_data = json.dump(report_data, file, ensure_ascii=False, indent=4)


class XMLReport(Report):
    def generate(self, headers: list, data: list, filename: str):
        root = ET.Element("Data")

        for row in data:
            row_elem = ET.SubElement(root, "Row")
            for i, header in enumerate(headers):
                col_elem = ET.SubElement(row_elem, "Column", name=header)
                col_elem.text = str(row[i])

        tree = ET.ElementTree(root)
        tree.write(filename, encoding="utf-8", xml_declaration=True)


headers = ["Заголовок 1", "Заголовок 2", "Заголовок 3"]
data = [[1, "a", True], [2, "b", False], [3, "c", True]]

excel_report = ExcelReport()
# report.generate(headers, data, "C:/Users/i.bondarenko/Desktop/222/report.xlsx")
csv_report = CSVReport()
# cvs_report.generate(headers, data, "C:/Users/i.bondarenko/Desktop/222/report.csv")
json_report = JSONReport()
# json_report.generate(headers, data, "C:/Users/i.bondarenko/Desktop/222/report.json")
xml_report = XMLReport()
# xml_report.generate(headers, data, "C:/Users/i.bondarenko/Desktop/222/report.xml")
